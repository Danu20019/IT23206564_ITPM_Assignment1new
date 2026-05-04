import openpyxl
from playwright.sync_api import sync_playwright
import time
import os
import sys

def run_automation():
    excel_file = "IT23206564_Assignment 1 - Test cases.xlsx"
    if not os.path.exists(excel_file):
        print(f"Excel file {excel_file} not found.", flush=True)
        return

    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active
    
    headers = [cell.value for cell in sheet[1]]
    input_col = headers.index("Input") + 1
    expected_col = headers.index("Expected Output") + 1
    actual_col = headers.index("Actual output") + 1
    status_col = headers.index("Status") + 1

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        print("Navigating...", flush=True)
        page.goto("https://www.pixelssuite.com/chat-translator", wait_until="networkidle")
        
        input_ta = page.locator('textarea[placeholder*="English text"]')
        output_ta = page.locator('textarea[placeholder*="Transliterated Sinhala"]')

        for row in range(2, sheet.max_row + 1):
            input_text = sheet.cell(row=row, column=input_col).value
            expected_output = sheet.cell(row=row, column=expected_col).value
            if not input_text: continue
                
            print(f"Row {row}: {input_text}", flush=True)
            
            input_ta.fill("")
            input_ta.fill(str(input_text))
            
            # Click by text using JS
            page.evaluate("""() => {
                const btns = Array.from(document.querySelectorAll('button'));
                const b = btns.find(x => x.innerText.includes('Transliterate'));
                if (b) b.click();
            }""")
            
            time.sleep(4)
            
            actual_output = output_ta.input_value()
            
            if not actual_output:
                # Try pressing enter
                input_ta.press("Enter")
                time.sleep(2)
                actual_output = output_ta.input_value()
            
            sheet.cell(row=row, column=actual_col).value = actual_output
            status = "PASS" if str(actual_output).strip() == str(expected_output).strip() else "FAIL"
            sheet.cell(row=row, column=status_col).value = status
            
            try: print(f"  Actual: {actual_output}", flush=True)
            except: print(f"  Actual: [Unicode]", flush=True)
            
        wb.save(excel_file)
        print("Done.", flush=True)
        browser.close()

if __name__ == "__main__":
    run_automation()
